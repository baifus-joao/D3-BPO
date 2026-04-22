from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from time import time
from types import SimpleNamespace

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import literal, or_, select, union_all
from sqlalchemy.orm import Session

from .erp import transaction_query
from .internal_finance_catalog import display_financial_category, predefined_subcategory_map
from .models import BankAccount, FinancialCategory, FinancialTransaction, PaymentMethod, Store


BRAND_DARK = "0E1F26"
BRAND_GRAPHITE = "2E2E2E"
BRAND_ACCENT = "22FFC4"
BRAND_TEXT_LIGHT = "ECF6F4"
BRAND_TEXT_DARK = "102821"
BRAND_MUTED = "8FA8A1"
BRAND_BORDER = "31444C"
REFERENCE_CACHE_TTL_SECONDS = 60
FORM_STATE_CACHE_TTL_SECONDS = 30
_cashflow_reference_cache: dict[str, object] = {"expires_at": 0.0, "value": None}
_cashflow_form_state_cache: dict[str, object] = {"expires_at": 0.0, "value": None}


def parse_date_input(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def safe_decimal(value: str | None) -> Decimal:
    text = str(value or "0").strip().replace("R$", "").replace(".", "").replace(",", ".")
    return Decimal(text or "0").copy_abs().quantize(Decimal("0.01"))


def signed_amount(item: FinancialTransaction) -> Decimal:
    amount = Decimal(item.amount)
    return amount if item.type == "ENTRADA" else -amount


def effective_date(item: FinancialTransaction) -> date:
    return item.realized_date or item.planned_date or item.transaction_date


def build_line_points(values: list[Decimal]) -> str:
    if not values:
        return "0,100 320,100"
    min_value = min(values)
    max_value = max(values)
    span = max_value - min_value or Decimal("1")
    points = []
    for index, value in enumerate(values):
        x = 10 + (300 * index / max(1, len(values) - 1))
        ratio = float((value - min_value) / span)
        y = 110 - (ratio * 90)
        points.append(f"{x:.2f},{y:.2f}")
    return " ".join(points)


def load_cashflow_reference_lists(db: Session) -> dict[str, list[object]]:
    now = time()
    if _cashflow_reference_cache["value"] is not None and float(_cashflow_reference_cache["expires_at"]) > now:
        return _cashflow_reference_cache["value"]  # type: ignore[return-value]

    references = union_all(
        select(literal("stores").label("kind"), Store.id.label("id"), Store.name.label("name")).where(Store.is_active.is_(True)),
        select(literal("accounts").label("kind"), BankAccount.id.label("id"), BankAccount.name.label("name")).where(BankAccount.is_active.is_(True)),
        select(literal("payment_methods").label("kind"), PaymentMethod.id.label("id"), PaymentMethod.name.label("name")).where(PaymentMethod.is_active.is_(True)),
    ).subquery()

    rows = db.execute(
        select(references.c.kind, references.c.id, references.c.name).order_by(references.c.kind, references.c.name)
    ).all()
    data = {"stores": [], "accounts": [], "categories": [], "payment_methods": []}
    for row in rows:
        data[row.kind].append(SimpleNamespace(id=row.id, name=row.name))
    categories = db.scalars(
        select(FinancialCategory)
        .where(FinancialCategory.is_active.is_(True))
        .order_by(FinancialCategory.type.asc(), FinancialCategory.name.asc())
    ).all()
    data["categories"] = [
        SimpleNamespace(
            id=item.id,
            name=display_financial_category(item.name),
            raw_name=item.name,
            type=item.type,
        )
        for item in categories
    ]

    _cashflow_reference_cache["expires_at"] = now + REFERENCE_CACHE_TTL_SECONDS
    _cashflow_reference_cache["value"] = data
    return data


def clear_cashflow_caches() -> None:
    _cashflow_reference_cache["expires_at"] = 0.0
    _cashflow_reference_cache["value"] = None
    _cashflow_form_state_cache["expires_at"] = 0.0
    _cashflow_form_state_cache["value"] = None


def _serialize_financial_row(item: FinancialTransaction, *, format_currency, format_short_date) -> dict[str, object]:
    installment_label = "-"
    if item.entry_mode == "parcelado" and item.installment_total:
        installment_label = f"{item.installment_number}/{item.installment_total}"
    elif item.entry_mode == "projecao" and item.projection_label:
        installment_label = item.projection_label
    return {
        "id": item.id,
        "date": format_short_date(item.transaction_date),
        "type": item.type,
        "description": item.description,
        "interested_party": item.interested_party or "-",
        "category": display_financial_category(item.category.name) if item.category else "-",
        "subcategory": item.subcategory or "-",
        "amount": format_currency(item.amount),
        "amount_value": Decimal(item.amount),
        "status": item.status,
        "status_class": "success" if item.status == "realizado" else "warning",
        "entry_mode": item.entry_mode or "avista",
        "entry_mode_label": "Parcelado" if item.entry_mode == "parcelado" else "Projecao" if item.entry_mode == "projecao" else "A vista",
        "installment_label": installment_label,
        "account": item.bank_account.name if item.bank_account else "-",
        "store": item.store.name if item.store else "-",
        "source": item.source,
        "planned_date": format_short_date(item.planned_date),
        "realized_date": format_short_date(item.realized_date),
        "editable": item.source != "conciliacao",
        "detail_url": f"/gestao/financeiro/lancamentos/{item.id}",
        "duplicate_url": f"/gestao/financeiro/novo?duplicate_id={item.id}",
    }


def apply_cashflow_filters(query, filters: dict[str, object]):
    if filters.get("source"):
        query = query.where(FinancialTransaction.source == filters["source"])
    if filters.get("date_from"):
        query = query.where(FinancialTransaction.transaction_date >= filters["date_from"])
    if filters.get("date_to"):
        query = query.where(FinancialTransaction.transaction_date <= filters["date_to"])
    if filters.get("category_id"):
        query = query.where(FinancialTransaction.category_id == filters["category_id"])
    if filters.get("store_id"):
        query = query.where(FinancialTransaction.store_id == filters["store_id"])
    if filters.get("status"):
        query = query.where(FinancialTransaction.status == filters["status"])
    if filters.get("account_id"):
        query = query.where(FinancialTransaction.bank_account_id == filters["account_id"])
    if filters.get("type"):
        query = query.where(FinancialTransaction.type == filters["type"])
    if filters.get("subcategory"):
        query = query.where(FinancialTransaction.subcategory.ilike(f"%{filters['subcategory']}%"))
    if filters.get("interested_party"):
        query = query.where(FinancialTransaction.interested_party.ilike(f"%{filters['interested_party']}%"))
    if filters.get("entry_mode"):
        query = query.where(FinancialTransaction.entry_mode == filters["entry_mode"])
    if filters.get("search"):
        term = f"%{filters['search']}%"
        query = query.where(
            or_(
                FinancialTransaction.description.ilike(term),
                FinancialTransaction.interested_party.ilike(term),
                FinancialTransaction.subcategory.ilike(term),
            )
        )
    return query


def load_cashflow_overview(
    db: Session,
    *,
    filters: dict[str, object],
    format_currency,
    format_short_date,
    page: int,
    page_size: int = 20,
) -> dict[str, object]:
    data_query = apply_cashflow_filters(
        select(
            FinancialTransaction.transaction_date,
            FinancialTransaction.type,
            FinancialTransaction.status,
            FinancialTransaction.amount,
            FinancialTransaction.planned_date,
            FinancialTransaction.realized_date,
            FinancialCategory.name.label("category_name"),
        ).outerjoin(FinancialCategory, FinancialCategory.id == FinancialTransaction.category_id),
        filters,
    )
    table_query = apply_cashflow_filters(
        transaction_query(db).order_by(FinancialTransaction.transaction_date.desc(), FinancialTransaction.id.desc()),
        filters,
    )

    transactions = db.scalars(table_query.offset((page - 1) * page_size).limit(page_size)).all()
    all_filtered = db.execute(data_query).all()

    today = date.today()
    realized_until_today = [
        item for item in all_filtered if item.status == "realizado" and (item.realized_date or item.planned_date or item.transaction_date) <= today
    ]
    saldo_atual = sum(
        (Decimal(item.amount) if item.type == "ENTRADA" else -Decimal(item.amount) for item in realized_until_today),
        Decimal("0"),
    )
    entradas = sum((Decimal(item.amount) for item in all_filtered if item.type == "ENTRADA"), Decimal("0"))
    saidas = sum((Decimal(item.amount) for item in all_filtered if item.type == "SAIDA"), Decimal("0"))
    resultado = entradas - saidas

    saldo_por_dia: dict[date, Decimal] = defaultdict(lambda: Decimal("0"))
    for item in sorted(all_filtered, key=lambda row: row.realized_date or row.planned_date or row.transaction_date):
        effective = item.realized_date or item.planned_date or item.transaction_date
        saldo_por_dia[effective] += Decimal(item.amount) if item.type == "ENTRADA" else -Decimal(item.amount)

    running = Decimal("0")
    line_values = []
    line_labels = []
    for day, value in sorted(saldo_por_dia.items()):
        running += value
        line_values.append(running)
        line_labels.append(day.strftime("%d/%m"))

    grouped_period: dict[str, dict[str, Decimal]] = defaultdict(lambda: {"ENTRADA": Decimal("0"), "SAIDA": Decimal("0")})
    for item in all_filtered:
        effective = item.realized_date or item.planned_date or item.transaction_date
        grouped_period[effective.strftime("%d/%m")][item.type] += Decimal(item.amount)
    bar_chart = [{"label": label, "entradas": values["ENTRADA"], "saidas": values["SAIDA"]} for label, values in list(grouped_period.items())[-8:]]
    max_bar_value = max([Decimal("1")] + [item["entradas"] for item in bar_chart] + [item["saidas"] for item in bar_chart])

    category_totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for item in all_filtered:
        label = display_financial_category(item.category_name or "Sem categoria")
        category_totals[label] += Decimal(item.amount)
    category_total_sum = sum(category_totals.values(), Decimal("0")) or Decimal("1")
    colors = ["#22ffc4", "#7aa2ff", "#ffd166", "#ff8a65", "#b48fff", "#4dd0e1"]
    cumulative = 0.0
    donut = []
    for index, (label, value) in enumerate(sorted(category_totals.items(), key=lambda pair: pair[1], reverse=True)[:6]):
        percentage = float(value / category_total_sum)
        donut.append(
            {
                "label": label,
                "value": format_currency(value),
                "percentage": percentage * 100,
                "dasharray": f"{percentage * 314:.2f} 314",
                "dashoffset": f"{-cumulative * 314:.2f}",
                "color": colors[index % len(colors)],
            }
        )
        cumulative += percentage

    rows = [
        _serialize_financial_row(item, format_currency=format_currency, format_short_date=format_short_date)
        for item in transactions
    ]

    return {
        "summary": {
            "saldo_atual": saldo_atual,
            "entradas_periodo": entradas,
            "saidas_periodo": saidas,
            "resultado_periodo": resultado,
        },
        "transactions": rows,
        "total_count": len(all_filtered),
        "has_more": len(all_filtered) > (page * page_size),
        "page": page,
        "line_chart": {"labels": line_labels, "points": build_line_points(line_values)},
        "bar_chart": [
            {**item, "entrada_pct": float((item["entradas"] / max_bar_value) * 100), "saida_pct": float((item["saidas"] / max_bar_value) * 100)}
            for item in bar_chart
        ],
        "donut_chart": donut,
    }


def build_cashflow_query(db: Session, *, filters: dict[str, object]):
    return apply_cashflow_filters(
        transaction_query(db).order_by(FinancialTransaction.transaction_date.desc(), FinancialTransaction.id.desc()),
        filters,
    )


def export_cashflow_workbook(
    db: Session,
    *,
    filters: dict[str, object],
    generated_by: str,
    generated_at: datetime | None = None,
) -> BytesIO:
    generated_at = generated_at or datetime.now()
    query = build_cashflow_query(db, filters=filters)
    all_filtered = db.scalars(query).all()

    today = date.today()
    realized_until_today = [item for item in all_filtered if item.status == "realizado" and effective_date(item) <= today]
    saldo_atual = sum((signed_amount(item) for item in realized_until_today), Decimal("0"))
    entradas = sum((Decimal(item.amount) for item in all_filtered if item.type == "ENTRADA"), Decimal("0"))
    saidas = sum((Decimal(item.amount) for item in all_filtered if item.type == "SAIDA"), Decimal("0"))
    resultado = entradas - saidas

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fluxo de Caixa"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A9"

    title_fill = PatternFill("solid", fgColor=BRAND_DARK)
    dark_fill = PatternFill("solid", fgColor=BRAND_GRAPHITE)
    muted_fill = PatternFill("solid", fgColor="172A33")
    header_fill = PatternFill("solid", fgColor="16323A")
    even_fill = PatternFill("solid", fgColor="132830")
    odd_fill = PatternFill("solid", fgColor="102028")
    thin_border = Border(
        left=Side(style="thin", color=BRAND_BORDER),
        right=Side(style="thin", color=BRAND_BORDER),
        top=Side(style="thin", color=BRAND_BORDER),
        bottom=Side(style="thin", color=BRAND_BORDER),
    )

    sheet.merge_cells("A1:J1")
    sheet["A1"] = "D3 GESTÃO | FLUXO DE CAIXA"
    sheet["A1"].fill = title_fill
    sheet["A1"].font = Font(color=BRAND_TEXT_LIGHT, bold=True, size=18)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet.merge_cells("A2:J2")
    sheet["A2"] = "Relatório interno gerado pelo sistema D3 Gestão"
    sheet["A2"].fill = dark_fill
    sheet["A2"].font = Font(color=BRAND_MUTED, size=11)
    sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    filter_labels = [
        ("Período", f"{filters['date_from'].strftime('%d/%m/%Y') if filters['date_from'] else 'Início'} até {filters['date_to'].strftime('%d/%m/%Y') if filters['date_to'] else 'Hoje'}"),
        ("Categoria", next((item.category.name for item in all_filtered if filters["category_id"] and item.category_id == filters["category_id"]), "Todas") if filters["category_id"] else "Todas"),
        ("Loja", next((item.store.name for item in all_filtered if filters["store_id"] and item.store_id == filters["store_id"]), "Todas") if filters["store_id"] else "Todas"),
        ("Conta", next((item.bank_account.name for item in all_filtered if filters["account_id"] and item.bank_account_id == filters["account_id"]), "Todas") if filters["account_id"] else "Todas"),
        ("Status", str(filters["status"] or "Todos").title()),
        ("Tipo", "Saída" if filters["type"] == "SAIDA" else "Entrada" if filters["type"] == "ENTRADA" else "Todos"),
        ("Gerado por", generated_by),
        ("Gerado em", generated_at.strftime("%d/%m/%Y %H:%M")),
    ]

    row = 4
    for index, (label, value) in enumerate(filter_labels):
        col = 1 if index % 2 == 0 else 6
        if col == 1 and index > 0 and index % 2 == 0:
            row += 1
        label_cell = sheet.cell(row=row, column=col, value=label)
        value_cell = sheet.cell(row=row, column=col + 1, value=value)
        sheet.merge_cells(start_row=row, start_column=col + 1, end_row=row, end_column=col + 4)
        label_cell.fill = muted_fill
        value_cell.fill = muted_fill
        label_cell.font = Font(color=BRAND_MUTED, bold=True, size=10)
        value_cell.font = Font(color=BRAND_TEXT_LIGHT, size=10)
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center")
        for c in range(col, col + 5):
            sheet.cell(row=row, column=c).border = thin_border

    summary_row = 8
    summary_blocks = [
        ("Saldo atual", saldo_atual, BRAND_DARK),
        ("Entradas", entradas, BRAND_ACCENT),
        ("Saídas", saidas, "FF8A65"),
        ("Resultado", resultado, "7AA2FF"),
    ]
    start_columns = [1, 3, 5, 7]
    for (label, value, color), start_col in zip(summary_blocks, start_columns, strict=False):
        sheet.merge_cells(start_row=summary_row, start_column=start_col, end_row=summary_row, end_column=start_col + 1)
        sheet.merge_cells(start_row=summary_row + 1, start_column=start_col, end_row=summary_row + 1, end_column=start_col + 1)
        label_cell = sheet.cell(summary_row, start_col, label)
        value_cell = sheet.cell(summary_row + 1, start_col, float(value))
        fill = PatternFill("solid", fgColor=color if color != BRAND_ACCENT else BRAND_ACCENT)
        label_cell.fill = fill
        value_cell.fill = fill
        label_cell.font = Font(color=BRAND_TEXT_DARK if color == BRAND_ACCENT else BRAND_TEXT_LIGHT, bold=True, size=10)
        value_cell.font = Font(color=BRAND_TEXT_DARK if color == BRAND_ACCENT else BRAND_TEXT_LIGHT, bold=True, size=14)
        label_cell.alignment = Alignment(horizontal="center")
        value_cell.alignment = Alignment(horizontal="center")
        value_cell.number_format = 'R$ #,##0.00'
        for merge_row in (summary_row, summary_row + 1):
            for c in range(start_col, start_col + 2):
                sheet.cell(merge_row, c).border = thin_border

    headers = [
        "Data",
        "Tipo",
        "Status",
        "Descrição",
        "Categoria",
        "Subcategoria",
        "Forma",
        "Conta",
        "Loja",
        "Valor",
    ]
    data_header_row = 11
    for idx, title in enumerate(headers, start=1):
        cell = sheet.cell(data_header_row, idx, title)
        cell.fill = header_fill
        cell.font = Font(color=BRAND_TEXT_LIGHT, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row_index, item in enumerate(all_filtered, start=data_header_row + 1):
        fill = even_fill if row_index % 2 == 0 else odd_fill
        values = [
            item.transaction_date,
            "Entrada" if item.type == "ENTRADA" else "Saída",
            item.status.title(),
            item.description,
            item.category.name if item.category else "",
            item.subcategory,
            item.payment_method.name if item.payment_method else "",
            item.bank_account.name if item.bank_account else "",
            item.store.name if item.store else "",
            float(item.amount),
        ]
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, col_index, value)
            cell.fill = fill
            cell.border = thin_border
            cell.font = Font(color=BRAND_TEXT_LIGHT, size=10)
            cell.alignment = Alignment(vertical="center")
            if col_index == 1 and value:
                cell.number_format = "DD/MM/YYYY"
            if col_index == 10:
                cell.number_format = 'R$ #,##0.00'
                if item.type == "SAIDA":
                    cell.font = Font(color="FF8A80", bold=True, size=10)
                else:
                    cell.font = Font(color=BRAND_ACCENT, bold=True, size=10)

    widths = {
        1: 14,
        2: 12,
        3: 14,
        4: 34,
        5: 22,
        6: 20,
        7: 18,
        8: 20,
        9: 18,
        10: 16,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_cashflow_form_state(db: Session) -> dict[str, object]:
    now = time()
    if _cashflow_form_state_cache["value"] is not None and float(_cashflow_form_state_cache["expires_at"]) > now:
        return _cashflow_form_state_cache["value"]  # type: ignore[return-value]

    recent_transactions = db.execute(
        select(
            FinancialTransaction.description,
            FinancialTransaction.category_id,
            FinancialTransaction.subcategory,
            FinancialCategory.name.label("category_name"),
        )
        .outerjoin(FinancialCategory, FinancialCategory.id == FinancialTransaction.category_id)
        .where(FinancialTransaction.source == "manual")
        .order_by(FinancialTransaction.created_at.desc(), FinancialTransaction.id.desc())
        .limit(200)
    ).all()

    subcategories_by_category: dict[str, list[str]] = defaultdict(list)
    description_suggestions: list[dict[str, object]] = []
    interested_suggestions: list[str] = []
    seen_descriptions: set[str] = set()
    seen_interested: set[str] = set()

    ranked_descriptions: dict[tuple[str, int | None, str], dict[str, object]] = {}
    for item in recent_transactions:
        category_name = item.category_name or ""
        if category_name and item.subcategory and item.subcategory not in subcategories_by_category[category_name]:
            subcategories_by_category[category_name].append(item.subcategory)

        normalized_description = item.description.strip()
        if not normalized_description:
            continue
        key = (normalized_description.lower(), item.category_id, item.subcategory.strip().lower())
        bucket = ranked_descriptions.setdefault(
            key,
            {
                "description": normalized_description,
                "category_id": item.category_id,
                "category_name": display_financial_category(category_name),
                "subcategory": item.subcategory.strip(),
                "usage_count": 0,
            },
        )
        bucket["usage_count"] += 1

    for item in sorted(
        ranked_descriptions.values(),
        key=lambda row: (-int(row["usage_count"]), str(row["description"]).lower()),
    ):
        description = str(item["description"])
        if description.lower() in seen_descriptions:
            continue
        seen_descriptions.add(description.lower())
        description_suggestions.append(item)
        if len(description_suggestions) >= 12:
            break

    interested_rows = db.scalars(
        select(FinancialTransaction.interested_party)
        .where(FinancialTransaction.interested_party != "")
        .order_by(FinancialTransaction.interested_party.asc())
        .limit(200)
    ).all()
    for item in interested_rows:
        normalized = str(item).strip()
        key = normalized.lower()
        if not normalized or key in seen_interested:
            continue
        seen_interested.add(key)
        interested_suggestions.append(normalized)
        if len(interested_suggestions) >= 30:
            break

    for category_name, subcategories in predefined_subcategory_map().items():
        for subcategory in subcategories:
            if subcategory not in subcategories_by_category[category_name]:
                subcategories_by_category[category_name].append(subcategory)

    value = {
        "today": date.today().strftime("%Y-%m-%d"),
        "subcategory_map": dict(subcategories_by_category),
        "description_suggestions": description_suggestions,
        "interested_suggestions": interested_suggestions,
    }
    _cashflow_form_state_cache["expires_at"] = now + FORM_STATE_CACHE_TTL_SECONDS
    _cashflow_form_state_cache["value"] = value
    return value


def load_internal_finance_overview(
    db: Session,
    *,
    filters: dict[str, object],
    format_currency,
    format_short_date,
    row_limit_per_report: int = 50,
) -> dict[str, object]:
    query = apply_cashflow_filters(
        transaction_query(db)
        .where(FinancialTransaction.source == "manual")
        .order_by(FinancialTransaction.transaction_date.desc(), FinancialTransaction.id.desc()),
        filters,
    )
    transactions = db.scalars(query).all()

    realized_until_today = [
        item
        for item in transactions
        if item.status == "realizado" and (item.realized_date or item.planned_date or item.transaction_date) <= date.today()
    ]
    saldo_atual = sum(
        (Decimal(item.amount) if item.type == "ENTRADA" else -Decimal(item.amount) for item in realized_until_today),
        Decimal("0"),
    )
    entradas = sum((Decimal(item.amount) for item in transactions if item.type == "ENTRADA"), Decimal("0"))
    saidas = sum((Decimal(item.amount) for item in transactions if item.type == "SAIDA"), Decimal("0"))

    all_payable_rows = [
        _serialize_financial_row(item, format_currency=format_currency, format_short_date=format_short_date)
        for item in transactions
        if item.type == "SAIDA"
    ]
    all_receivable_rows = [
        _serialize_financial_row(item, format_currency=format_currency, format_short_date=format_short_date)
        for item in transactions
        if item.type == "ENTRADA"
    ]
    payable_rows = all_payable_rows[:row_limit_per_report]
    receivable_rows = all_receivable_rows[:row_limit_per_report]

    return {
        "summary": {
            "saldo_atual": saldo_atual,
            "entradas_periodo": entradas,
            "saidas_periodo": saidas,
            "resultado_periodo": entradas - saidas,
        },
        "payables_report": {
            "rows": payable_rows,
            "count": len(all_payable_rows),
            "total": sum((item["amount_value"] for item in all_payable_rows), Decimal("0")),
            "pending_count": len([item for item in all_payable_rows if item["status"] == "previsto"]),
            "has_more": len(all_payable_rows) > len(payable_rows),
        },
        "receivables_report": {
            "rows": receivable_rows,
            "count": len(all_receivable_rows),
            "total": sum((item["amount_value"] for item in all_receivable_rows), Decimal("0")),
            "pending_count": len([item for item in all_receivable_rows if item["status"] == "previsto"]),
            "has_more": len(all_receivable_rows) > len(receivable_rows),
        },
    }
