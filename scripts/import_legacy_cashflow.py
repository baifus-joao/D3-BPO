from __future__ import annotations

import argparse
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from webapp.db import SessionLocal
from webapp.models import FinancialCategory, FinancialTransaction, PaymentMethod, User


LEGACY_HEADER = (
    "Data pgto",
    "Descrição",
    "Tipo (Entrada/Saída)",
    "Categoria",
    "Forma de Pagamento",
    "Valor(R$)",
)

ENTRY_COLOR = "#22ffc4"
EXIT_COLOR = "#ff8a65"


def normalize_text(value: object) -> str:
    return str(value or "").strip()


def normalize_type(value: str) -> str:
    text = normalize_text(value).lower()
    if text == "entrada":
        return "ENTRADA"
    if text == "saída" or text == "saida":
        return "SAIDA"
    raise ValueError(f"Tipo inválido no legado: {value!r}")


def normalize_decimal(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    text = normalize_text(value).replace("R$", "").replace(".", "").replace(",", ".")
    return Decimal(text).quantize(Decimal("0.01"))


def normalize_date(value: object):
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%d/%m/%Y").date()
    raise ValueError(f"Data inválida no legado: {value!r}")


def build_payment_code(name: str) -> str:
    cleaned = normalize_text(name).upper()
    replacements = {
        "Ã": "A",
        "Á": "A",
        "À": "A",
        "Â": "A",
        "É": "E",
        "Ê": "E",
        "Í": "I",
        "Ó": "O",
        "Ô": "O",
        "Õ": "O",
        "Ú": "U",
        "Ç": "C",
        " ": "_",
        "/": "_",
        "-": "_",
    }
    for source, target in replacements.items():
        cleaned = cleaned.replace(source, target)
    return "".join(ch for ch in cleaned if ch.isalnum() or ch == "_")[:40] or "LEGADO"


def find_header_row(worksheet) -> int:
    for row_index in range(1, worksheet.max_row + 1):
        values = tuple(worksheet.cell(row_index, col).value for col in range(1, 7))
        if values == LEGACY_HEADER:
            return row_index
    raise ValueError(f"Cabeçalho legado não encontrado na aba {worksheet.title!r}")


def get_or_create_category(db, name: str, type_value: str) -> FinancialCategory:
    category = db.scalar(select(FinancialCategory).where(FinancialCategory.name == name))
    if category:
        if category.type != type_value:
            raise ValueError(f"Categoria existente com tipo diferente: {name!r}")
        return category
    color = ENTRY_COLOR if type_value == "ENTRADA" else EXIT_COLOR
    category = FinancialCategory(name=name, type=type_value, color=color, is_active=True)
    db.add(category)
    db.flush()
    return category


def get_or_create_payment_method(db, name: str) -> PaymentMethod | None:
    normalized = normalize_text(name)
    if not normalized:
        return None
    method = db.scalar(select(PaymentMethod).where(PaymentMethod.name == normalized))
    if method:
        return method
    code = build_payment_code(normalized)
    method = db.scalar(select(PaymentMethod).where(PaymentMethod.code == code))
    if method:
        return method
    method = PaymentMethod(name=normalized, code=code, is_active=True)
    db.add(method)
    db.flush()
    return method


def load_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook["Lançamentos"]
    header_row = find_header_row(sheet)
    rows: list[dict[str, object]] = []
    for row_index in range(header_row + 1, sheet.max_row + 1):
        date_value = sheet.cell(row_index, 1).value
        description = sheet.cell(row_index, 2).value
        type_value = sheet.cell(row_index, 3).value
        category = sheet.cell(row_index, 4).value
        payment_method = sheet.cell(row_index, 5).value
        amount = sheet.cell(row_index, 6).value
        if not any(value not in (None, "") for value in (date_value, description, type_value, category, payment_method, amount)):
            continue
        if not hasattr(date_value, "date") and not (isinstance(date_value, str) and "/" in date_value):
            continue
        rows.append(
            {
                "transaction_date": normalize_date(date_value),
                "description": normalize_text(description),
                "type": normalize_type(str(type_value)),
                "category_name": normalize_text(category),
                "payment_method_name": normalize_text(payment_method),
                "amount": normalize_decimal(amount),
            }
        )
    return rows


def import_legacy_cashflow(path: Path, created_by_email: str | None, source_reference: str, dry_run: bool) -> None:
    rows = load_rows(path)
    with SessionLocal() as db:
        created_by_id = None
        if created_by_email:
            user = db.scalar(select(User).where(User.email == created_by_email.strip().lower()))
            if not user:
                raise ValueError(f"Usuário não encontrado: {created_by_email}")
            created_by_id = user.id

        inserted = 0
        created_categories: set[str] = set()
        created_methods: set[str] = set()

        for item in rows:
            category = get_or_create_category(db, str(item["category_name"]), str(item["type"]))
            if category.name == item["category_name"]:
                created_categories.add(category.name)
            method = get_or_create_payment_method(db, str(item["payment_method_name"]))
            if method and method.name == item["payment_method_name"]:
                created_methods.add(method.name)

            duplicate = db.scalar(
                select(FinancialTransaction).where(
                    FinancialTransaction.transaction_date == item["transaction_date"],
                    FinancialTransaction.description == item["description"],
                    FinancialTransaction.amount == item["amount"],
                    FinancialTransaction.type == item["type"],
                )
            )
            if duplicate:
                continue

            db.add(
                FinancialTransaction(
                    transaction_date=item["transaction_date"],
                    type=str(item["type"]),
                    description=str(item["description"]),
                    category_id=category.id,
                    subcategory="",
                    amount=item["amount"],
                    payment_method_id=method.id if method else None,
                    bank_account_id=None,
                    store_id=None,
                    source="manual",
                    source_reference=source_reference,
                    status="realizado",
                    planned_date=item["transaction_date"],
                    realized_date=item["transaction_date"],
                    created_by_user_id=created_by_id,
                )
            )
            inserted += 1

        if dry_run:
            db.rollback()
        else:
            db.commit()

    mode_label = "DRY RUN" if dry_run else "IMPORTADO"
    print(mode_label)
    print(f"arquivo={path}")
    print(f"linhas_lidas={len(rows)}")
    print(f"lancamentos_novos={inserted}")
    print(f"source_reference={source_reference}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa o fluxo de caixa do layout legado para d3_financial_transactions.")
    parser.add_argument("file", help="Caminho do arquivo .xlsx legado")
    parser.add_argument("--created-by-email", help="E-mail do usuário a registrar como criador")
    parser.add_argument("--source-reference", default="import-legado-planilha", help="Identificador de origem para os lançamentos importados")
    parser.add_argument("--dry-run", action="store_true", help="Lê e valida, mas não grava no banco")
    args = parser.parse_args()

    import_legacy_cashflow(
        path=Path(args.file).resolve(),
        created_by_email=args.created_by_email,
        source_reference=args.source_reference,
        dry_run=args.dry_run,
    )


if __name__ == "__main__":
    main()
