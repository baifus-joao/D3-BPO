from pathlib import Path

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _auto_width(ws):
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)


def _format_sheet(ws, table_name, date_cols=None, currency_cols=None, int_cols=None):
    date_cols = set(date_cols or [])
    currency_cols = set(currency_cols or [])
    int_cols = set(int_cols or [])

    ws.freeze_panes = "A2"
    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.auto_filter.ref = ref

        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    idx_by_name = {ws.cell(row=1, column=i).value: i for i in range(1, ws.max_column + 1)}

    for col_name in date_cols:
        if col_name in idx_by_name:
            col = get_column_letter(idx_by_name[col_name])
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = "DD/MM/YYYY"
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

    for col_name in currency_cols:
        if col_name in idx_by_name:
            col = get_column_letter(idx_by_name[col_name])
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = "R$ #,##0.00"

    for col_name in int_cols:
        if col_name in idx_by_name:
            col = get_column_letter(idx_by_name[col_name])
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

    _auto_width(ws)


def _add_daily_chart(ws):
    if ws.max_row < 2:
        return
    chart = LineChart()
    chart.title = "Valores Recebidos por Dia"
    chart.style = 2
    chart.y_axis.title = "Valor (R$)"
    chart.x_axis.title = "Data"
    chart.height = 8
    chart.width = 16

    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=3, max_col=4, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend.position = "r"
    ws.add_chart(chart, "I2")


def _add_forecast_chart(ws):
    if ws.max_row < 2:
        return
    chart = LineChart()
    chart.title = "Previsao de Recebimentos por Data"
    chart.style = 2
    chart.y_axis.title = "Valor (R$)"
    chart.x_axis.title = "Data"
    chart.height = 8
    chart.width = 16

    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=3, max_col=4, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend.position = "r"
    ws.add_chart(chart, "I2")


def _build_summary_sheet(
    writer,
    output_path,
    vendas_original,
    receb_original,
    received_daily,
    forecast_sales,
    paid_missing,
):
    generated_at = pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
    summary = pd.DataFrame(
        [
            {"Secao": "Arquivo gerado", "Detalhe": output_path.name},
            {"Secao": "Data de geracao", "Detalhe": generated_at},
            {"Secao": "Linhas no relatorio de vendas", "Detalhe": len(vendas_original)},
            {"Secao": "Linhas no relatorio de recebimentos", "Detalhe": len(receb_original)},
            {"Secao": "Dias com recebimento", "Detalhe": len(received_daily)},
            {"Secao": "Datas previstas de recebimento", "Detalhe": len(forecast_sales)},
            {"Secao": "Vendas pagas sem recebimento", "Detalhe": len(paid_missing)},
            {"Secao": "Como ler", "Detalhe": "1. Confira esta aba para um resumo geral."},
            {"Secao": "Como ler", "Detalhe": "2. Use 'Vendas Pagas sem Receb' para localizar pendencias."},
            {"Secao": "Como ler", "Detalhe": "3. Use as abas originais para auditoria e conferencia."},
        ]
    )
    summary.to_excel(writer, sheet_name="Resumo", index=False)


def save_objective_report(
    vendas_original,
    receb_original,
    received_daily,
    forecast_sales,
    paid_missing,
    output_path,
):
    output_path = Path(output_path)

    target = output_path
    try:
        open(target, "a").close()
    except PermissionError:
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        target = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        _build_summary_sheet(
            writer,
            target,
            vendas_original,
            receb_original,
            received_daily,
            forecast_sales,
            paid_missing,
        )
        vendas_original.to_excel(writer, sheet_name="Relatorio Vendas (Orig)", index=False)
        receb_original.to_excel(writer, sheet_name="Relatorio Receb (Orig)", index=False)
        received_daily.to_excel(writer, sheet_name="Recebido_por_Dia", index=False)
        forecast_sales.to_excel(writer, sheet_name="Previsao Receb Vendas", index=False)
        paid_missing.to_excel(writer, sheet_name="Vendas Pagas sem Receb", index=False)

        _format_sheet(writer.sheets["Resumo"], "TblResumo")
        _format_sheet(writer.sheets["Relatorio Vendas (Orig)"], "TblVendasOrig")
        _format_sheet(writer.sheets["Relatorio Receb (Orig)"], "TblRecebOrig")
        _format_sheet(
            writer.sheets["Recebido_por_Dia"],
            "TblRecebidoDia",
            date_cols={"Data"},
            currency_cols={"Valor_Bruto_Recebido", "Valor_Liquido_Recebido", "Desconto_MDR"},
            int_cols={"Quantidade_Lancamentos", "Ano", "Mes"},
        )
        _format_sheet(
            writer.sheets["Previsao Receb Vendas"],
            "TblPrevRecebVendas",
            date_cols={"Data"},
            currency_cols={"Valor_Bruto_Previsto", "Valor_Liquido_Previsto"},
            int_cols={"Quantidade_Parcelas", "Ano", "Mes"},
        )
        _format_sheet(
            writer.sheets["Vendas Pagas sem Receb"],
            "TblPagasSemReceb",
            date_cols={"Data Venda", "Data Prevista"},
            currency_cols={"Valor Parcela Venda", "Valor Liquido Venda"},
        )

        _add_daily_chart(writer.sheets["Recebido_por_Dia"])
        _add_forecast_chart(writer.sheets["Previsao Receb Vendas"])

    return target.resolve()
